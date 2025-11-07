# data/data_loader.py
from datetime import datetime
import warnings
import pandas as pd
from openpyxl import load_workbook
import calendar
import os

from data.connectors.cdf_connector import CDFConnector
from data.connectors.sql_connector import SQLConnector
from utils.dates import normalize_month_names, get_all_months, calculate_duration
from utils.file_manager import get_capex_config_path

class DataLoader:
    def __init__(self):
        # Inicializar conectores
        self.sql_connector = SQLConnector()
        self.sql_connector.connect() # Aquí puedes asignar un conector SQL real
        self.cdf_connector = CDFConnector()
        self.cdf_connector.connect()
        self._budget_data = None
        self._cdf_cache = None
        self.DIAS_MOVILIZACION = 1

    
    # ---------------------------------------------------
    # Métodos SQL
    # ---------------------------------------------------
    def load_from_sql(self, query):
        if self.sql_connector is None:
            raise ValueError("SQL Connector is not initialized")
        return self.sql_connector.fetch_data(query)
    


    def load_rig_rates(self):
        """
        Carga las tarifas de rig (R2, R10, etc.) desde la tabla VT_RIG_OFFER_en_US
        para ciertos rigs específicos.
        """
        query = """
        SELECT 
            START_DATETIME,
            ITEM_NAME AS RIG,
            R2 AS daily_operating_rate_hr,
            R10 AS standby_rate_crew_hr,
            R6 AS certification_rate,
            R48 AS ambulance_day,
            R49 AS same_pad,
            R21 AS rig_move_0_10km,
            R50 AS rig_move_10_20km,
            R51 AS rig_move_20_25km,
            R52 AS rig_move_25_35km,
            R53 AS rig_move_35_50km,
            R24 AS rig_move_50_75km,
            R25 AS rig_move_75_100km,
            R27 AS rig_move_125_150km,
            R54 AS extras_per_job,
            R55 AS extras_per_hr
        FROM VT_RIG_OFFER_en_US
        WHERE ITEM_NAME IN ('TUSCANY - 111','SINOPEC - 932','ORIENDRILL 903')
        """
        df = self.load_from_sql(query)
        return df
    
    def fetch_fails_by_year(self, year):
        """
        Conecta a la base de datos y obtiene la suma de fallas (FAILS) agrupadas por mes para el año indicado.
        
        Args:
            year (int): El año para el cual se desea obtener el total de fallas.
        
        Returns:
            pd.DataFrame: DataFrame con columnas 'Month' y 'TotalFails'.
        """
        start_date = f"01/01/{year}"
        end_date = f"12/31/{year}"
        
        query = f"""
        DECLARE @START_DATETIME DATETIME = '{start_date}', 
                @END_DATETIME DATETIME = '{end_date}';

        SELECT 
            MONTH(D.DATETIME) AS Month,
            SUM(CAST(ISNULL(V.LEV1, 0) AS FLOAT)) AS TotalFails
        FROM DATE_INFO D
        LEFT JOIN dbo.V_FAILS_ELEMENT_STATISTICS V
            ON V.MENSUAL = D.DATETIME
        WHERE D.DATE_TYPE = 'M'
        AND D.DATETIME BETWEEN @START_DATETIME AND @END_DATETIME
        GROUP BY MONTH(D.DATETIME)
        ORDER BY Month;
        """
        
        if not self.sql_connector:
            from data.connectors.sql_connector import SQLConnector
            from config import DB_CONFIG
            self.sql_connector = SQLConnector(DB_CONFIG)
            self.sql_connector.connect()

        df = self.sql_connector.fetch_data(query)
        return df
    # ---------------------------------------------------
    # Calcular duración promedio entre START_WO y END_WO
    # ---------------------------------------------------
    def calcular_duracion_movilizacion(self):
        query = """
            SELECT ITEM_NAME, START_WO, END_WO, START_RIG_MOV, END_RIG_MOV
            FROM VT_WELLJOBLOG_en_US
        """
        df = self.load_from_sql(query)
        print("BOENAS")
        print(list(df.columns))
        df['START_WO'] = pd.to_datetime(df['START_WO'], errors='coerce')
        df['END_WO'] = pd.to_datetime(df['END_WO'], errors='coerce')
        df['START_RIG_MOV'] = pd.to_datetime(df['START_RIG_MOV'], errors='coerce')
        df['END_RIG_MOV'] = pd.to_datetime(df['END_RIG_MOV'], errors='coerce')
        # Calcular diferencia en días
        df['DURACION_DIAS_MOVILIZACION'] = (df['END_RIG_MOV'] - df['START_RIG_MOV']).dt.total_seconds() / 86400
        df_result = df[["ITEM_NAME", "DURACION_DIAS_MOVILIZACION"]].dropna()
        
        return df_result

    def obtener_pozos_services(self):
        query = """
            SELECT WELL, p_WELLBORE FROM HIERARCHY(GETDATE())
        """
        df = self.load_from_sql(query)
        print("COLUMNAS DE LOS POZOS DE SERVICES")
        print(df)

    
    #def calcular_duracion_promedio(self):
        #query = """
        #SELECT ITEM_NAME, START_WO, END_WO
        #FROM VT_WELLJOBLOG_en_US
        #WHERE YEAR(END_WO) = 2025
        #AND PLAN_TYPE_TEXT = 'Opex'
        #"""
        """
        df = self.load_from_sql(query)
        print(df)
        print(self.calcular_duracion_movilizacion())
        print(self.obtener_pozos_services())
        # Asegurar que sean tipo datetime
        df['START_WO'] = pd.to_datetime(df['START_WO'], errors='coerce')
        df['END_WO'] = pd.to_datetime(df['END_WO'], errors='coerce')

        # Calcular diferencia en días
        df['DURACION_DIAS'] = (df['END_WO'] - df['START_WO']).dt.total_seconds() / 86400

        # Sumar 2 días adicionales
        df['DURACION_DIAS'] = df['DURACION_DIAS'] + self.DIAS_MOVILIZACION

        #print(df['DURACION_DIAS'])

        # Filtrar duraciones válidas
        duraciones_validas = df['DURACION_DIAS'].dropna()
        duraciones_validas = duraciones_validas[duraciones_validas >= 0]

        df_result = df[["ITEM_NAME", "DURACION_DIAS"]].dropna()
        df_result = df_result[df_result["DURACION_DIAS"] >= 0]
        print("HOLAAA")
        print(df_result)
        return df_result
        """
        
    def calcular_duracion_promedio(self):
        query = """
        SELECT ITEM_NAME, START_WO, END_WO, START_SUSPEN, END_SUSPEN
        FROM VT_WELLJOBLOG_en_US
        WHERE YEAR(END_WO) = 2025
        AND PLAN_TYPE_TEXT = 'Opex'
        """
        df = self.load_from_sql(query)
        print(df)
        print(self.calcular_duracion_movilizacion())
        print(self.obtener_pozos_services())

        # Asegurar que sean tipo datetime
        for col in ['START_WO', 'END_WO', 'START_SUSPEN', 'END_SUSPEN']:
            df[col] = pd.to_datetime(df[col], errors='coerce')

        # Calcular duración según la lógica SQL
        def calcular_duracion(row):
            if pd.isna(row['START_SUSPEN']):
                return (row['END_WO'] - row['START_WO']).days
            elif pd.isna(row['END_SUSPEN']):
                return (row['START_SUSPEN'] - row['START_WO']).days
            else:
                return (row['START_SUSPEN'] - row['START_WO']).days + (row['END_WO'] - row['END_SUSPEN']).days

        df['DURACION_DIAS'] = df.apply(calcular_duracion, axis=1)

        # Sumar días de movilización
        df['DURACION_DIAS'] = df['DURACION_DIAS']+1 #+ self.DIAS_MOVILIZACION se mantiene uno por falta de datos en Avocet

        # Filtrar duraciones válidas
        duraciones_validas = df['DURACION_DIAS'].dropna()
        duraciones_validas = duraciones_validas[duraciones_validas >= 0]

        df_result = df[["ITEM_NAME", "DURACION_DIAS"]].dropna()
        df_result = df_result[df_result["DURACION_DIAS"] >= 0]

        print("HOLAAA")
        print(df_result)
        return df_result

    # ---------------------------------------------------
    # Métodos para budget
    # ---------------------------------------------------
    
    def load_budget_for_line(self, year: int, line_col: str) -> pd.DataFrame:
        """
        1) Llama a self.load_budget_data(year)
        2) Filtra y agrupa por MONTH
        3) Renombra la col => "Budget"
        4) Devuelve DF
        """
        budget_df = self.load_budget_data(year=year)
        if line_col not in budget_df.columns:
            return pd.DataFrame({'MONTH':[], 'Budget':[]})
        grouped = budget_df.groupby('MONTH')[line_col].sum().reset_index()
        grouped.rename(columns={line_col:'Budget'}, inplace=True)
        # normalizar mes
        # grouped['MONTH'] = normalize_month_names(grouped['MONTH'])
        # agrupar por mes si quieres un 2do group_by
        # ...
        return grouped
    
    def get_capex_yes_month_indices(self):
        """
        Lee el archivo de configuración de meses CAPEX y devuelve una lista 
        con los números de los meses (1-12) marcados como "Yes".
        """
        config_path = get_capex_config_path()
        months = get_all_months()
        
        # Si el archivo no existe, retorna una lista vacía.
        if not os.path.exists(config_path):
            return []
        try:
            df = pd.read_csv(config_path)
            yes_months_df = df[df['Capex'] == 'Yes']
            yes_month_names = yes_months_df['Month'].tolist()
            indices = [months.index(name) + 1 for name in yes_month_names]
            return indices
            
        except Exception as e:
            raise Exception(f"⚠️ Error al leer el archivo de configuración CAPEX: {e}")

    
    def load_budget_data_all_years(self, start_year=2023, end_year=2025, sheet_name="Wells", table_name="Table1"):
        """
        Carga y procesa el reporte de presupuesto para todos los años (por ejemplo, de 2019 a 2025)
        leyendo una tabla específica del Excel, y lo cachea en self._budget_data.
        
        Si ya se cargó, retorna el DataFrame cacheado.
        
        Args:
            start_year (int): Año inicial.
            end_year (int): Año final.
            sheet_name (str): Nombre de la hoja donde se encuentra la tabla.
            table_name (str): Nombre de la tabla definida en la hoja.
            
        Returns:
            pd.DataFrame: DataFrame con los datos filtrados para los años indicados.
        """
        if self._budget_data is not None:
            return self._budget_data

        from utils.file_manager import obtener_archivo_reporte_actual  # Asumiendo que definiste load_table_from_excel

        archivo_excel = obtener_archivo_reporte_actual()

        if archivo_excel:
            print(f"Procesando archivo de presupuesto: {archivo_excel}")
            try:
                df = self.load_table_from_excel(archivo_excel, sheet_name, table_name)
            except ValueError as e:
                print(e)
                self._budget_data = pd.DataFrame()
                return self._budget_data

            # Convertir a DataFrame (por si load_table_from_excel ya lo devuelve, esto es opcional)
            df = pd.DataFrame(df)


            
            # Filtrar por el rango de años
            df = df[(df["YEAR"] >= start_year) & (df["YEAR"] <= end_year)]
            self._budget_data = df  # Guardar en cache
            return df
        else:
            print("No se encontró ningún archivo Excel para el presupuesto.")
            self._budget_data = pd.DataFrame()
            return self._budget_data


    def load_budget_data_per_year(self, year=2025):
        """
        Filtra el DataFrame cacheado para retornar solo los datos del año especificado.
        
        Args:
            year (int): Año para el cual se desea el presupuesto.
            
        Returns:
            pd.DataFrame: DataFrame filtrado para ese año, con la columna MONTH normalizada.
        """
        df = self.load_budget_data_all_years()
        if df.empty:
            return df


        df_year = df[df['YEAR'] == year].copy()
        df_year['MONTH'] = normalize_month_names(df_year['MONTH'])
        df_year = df_year[df_year['ACTIVITY'] == 'WO OPEX']

        return df_year
    
    def load_budget_data_from_excel(self, file_path, sheet_name, year=2025):
        """
        Carga y filtra los datos de presupuesto desde un Excel para un año dado.
        """
        try:
            data = pd.read_excel(file_path, sheet_name=sheet_name)
            data = data[data['YEAR'] == year]
            print("Budget data:")
            print(data.head(5))
            return data
        except Exception as e:
            raise ValueError(f"Error al procesar el archivo Excel: {e}")
        
    def load_budget_for_line(self, year: int, line_col: str) -> pd.DataFrame:
        """
        1) Llama a self.load_budget_data(year)
        2) Filtra y agrupa por MONTH
        3) Renombra la col => "ACTUAL_COST"
        4) Devuelve DF
        """
        budget_df = self.load_budget_data_per_year(year=year)
        if line_col not in budget_df.columns:
            return pd.DataFrame({'MONTH':[], 'ACTUAL_COST':[]})
        grouped = budget_df.groupby('MONTH')[line_col].sum().reset_index()
        grouped.rename(columns={line_col:'ACTUAL_COST'}, inplace=True)
        # normalizar mes
        # grouped['MONTH'] = normalize_month_names(grouped['MONTH'])
        # agrupar por mes si quieres un 2do group_by
        # ...
        return grouped


    # ---------------------------------------------------
    # CDF 
    # ---------------------------------------------------

    def load_cdf_activities(self, data_loader, year: int, c1_min=1, c1_max=17) -> pd.DataFrame:
        """
        1) Carga el DataFrame de CDF
        2) Aplica calculate_duration
        3) Filtra por 'End' en el año deseado
        4) Crea month_num, activity_type
        5) Filtra activity_type in [C1.x]
        6) Retorna DF final
        """
        df = data_loader.load_from_cognite()
        df = calculate_duration(df)
        df = df[df['End'].dt.year == year].copy()
        df['month_num'] = df['End'].dt.month
        df['activity_type'] = df['activity_type'].fillna('')
        # Filtrar c1.x
        c1_list = [f"C1.{i}" for i in range(c1_min, c1_max)]
        df = df[df['activity_type'].isin(c1_list)]
    
        #df = pd.read_excel('cdf_Data.xlsx')
        #df.to_excel('cdf_Data.xlsx')
        print(df)
        return df
    
    def load_from_cognite(self, database='jobs_catalogue', table='jobs_catalogue', limit=None):
        """
        Carga datos de Cognite usando el conector (CDFConnector) ya creado.
        Si ya se cargaron previamente, devuelve el cache.
        """
        if self._cdf_cache is not None:
            # Devolver datos cacheados
            return self._cdf_cache

        # Si aún no se han cargado, se realiza la consulta
        if not hasattr(self, 'cdf_connector') or self.cdf_connector is None:
            from data.connectors.cdf_connector import CDFConnector
            from config import COGNITE_CONFIG
            self.cdf_connector = CDFConnector(COGNITE_CONFIG)
            self.cdf_connector.connect()

        query = {'database': database, 'table': table, 'limit': limit}
        try:
            df = self.cdf_connector.fetch_data(query)
            # Convertir la columna 'Start' a datetime si es necesario
            if 'Start' in df.columns:
                df['Start'] = pd.to_datetime(df['Start'], errors='coerce')
            # Cachear el DataFrame para la sesión actual
            self._cdf_cache = df
            print("Problema aqui")
            print(df)
            return df
        except Exception as e:
            print(f"Error al obtener datos de Cognite: {e}")
            return pd.DataFrame()


    def group_cdf_by_month(self, df: pd.DataFrame, col='activity_type') -> pd.DataFrame:
        """
        Agrupa las actividades en df por 'month_num' y devuelve un DataFrame con
        month_num y la lista de activity_type en cada mes.
        """
        grouped = df.groupby('month_num')[col].agg(list).reset_index()
        return grouped


    def clear_cdf_cache(self):
        """
        Método para limpiar la caché de datos de Cognite, en caso de querer refrescar la información.
        """
        self._cdf_cache = None
        

    def fetch_capex_activities_for_year(self, year=None):
        """
        1) Obtiene de Cognite Data Fusion todas las actividades.
        2) Aplica 'calculate_duration' para obtener duración entre Start y End.
        3) Filtra por actividades CAPEX (activity_type in 'W1'..'W12').
        4) Filtra por actividades cuya fecha de término ('End') ocurra en el año indicado.
        5) Retorna un DataFrame con las columnas necesarias.
        """
        if not self.cdf_connector.client:
            print("❌ No hay conexión activa a Cognite (CDF).")
            return pd.DataFrame()

        if year is None:
            year = datetime.now().year

        query = {
            "database": "jobs_catalogue",
            "table": "jobs_catalogue",
            "limit": None
        }
        df = self.cdf_connector.fetch_data(query)
        if df.empty:
            print("⚠️ No se obtuvieron datos desde CDF (jobs_catalogue).")
            return pd.DataFrame()

        # Calcular duración y verificar campos necesarios
        df = calculate_duration(df)

        if 'End' not in df.columns or not pd.api.types.is_datetime64_any_dtype(df['End']):
            print("⚠️ La columna 'End' no está disponible o no es una fecha.")
            return pd.DataFrame()

        if 'activity_type' not in df.columns:
            print("⚠️ No se encontró la columna 'activity_type'.")
            return pd.DataFrame()

        # Filtrar por End en el año deseado
        df = df[df['End'].dt.year == year]
        print(f"📅 Actividades con 'End' en {year}: {len(df)}")

        # Filtrar por tipos de actividad CAPEX (W1..W12)
        capex_types = [f"W{i}" for i in range(1, 13)]
        df_capex = df[df['activity_type'].isin(capex_types)]

        if df_capex.empty:
            print(f"ℹ️ No hay actividades CAPEX (W1..W12) finalizadas en {year}.")
            return pd.DataFrame()

        return df_capex


    def process_capex_data(self, capex_df):
        """
        1) Convierte la columna 'Start' a datetime (si no está ya).
        2) Convierte 'duration' a numérico (horas).
        3) Agrega 48 horas de buffer a 'duration'.
        4) Calcula 'End' = 'Start' + 'duration' (en horas).
        Retorna un DataFrame listo para distribuir días por mes.
        """
        df = capex_df.copy()

        # Asegurar que 'Start' sea datetime
        df['Start'] = pd.to_datetime(df['Start'], errors='coerce')
        if df['Start'].isnull().any():
            print("Advertencia: Hay valores no válidos en la columna 'Start'. Se descartan esas filas.")
            df.dropna(subset=['Start'], inplace=True)

        # Asegurar que 'duration' sea numérico
        df['duration'] = pd.to_numeric(df['duration'], errors='coerce').fillna(0)

        # Sumar buffer de 48 horas
        df['duration'] += 48

        # Calcular 'End'
        df['End'] = df['Start'] + pd.to_timedelta(df['duration'], unit='h')

        return df

    def distribute_days_by_month(self, capex_data):
        """
        Distribuye y agrupa los días de CAPEX por mes (ej. 'January', 'February', ...).
        - Recorre cada rango (Start, End).
        - Para cada mes en el que cae la actividad, suma los días correspondientes.
        Retorna un dict {'January': X, 'February': Y, ...} o un DataFrame, según prefieras.
        """
        capex_monthly = {}

        for _, row in capex_data.iterrows():
            start_date = row['Start']
            end_date = row['End']

            if pd.isnull(start_date) or pd.isnull(end_date):
                continue
            if end_date < start_date:
                continue

            current_date = start_date

            while current_date <= end_date:
                # Mes numérico
                mes_num = current_date.month
                # Nombre del mes en inglés (January..December)
                mes_name = calendar.month_name[mes_num]

                # Días totales en el mes actual
                month_days = calendar.monthrange(current_date.year, mes_num)[1]

                # Días restantes para terminar la actividad
                remaining_in_activity = (end_date - current_date).days + 1
                # Días restantes para terminar el mes
                remaining_in_month = (month_days - current_date.day + 1)

                days_in_month = min(remaining_in_activity, remaining_in_month)

                if mes_name not in capex_monthly:
                    capex_monthly[mes_name] = 0
                capex_monthly[mes_name] += days_in_month

                # Avanzar al 1er día del siguiente mes
                next_month = current_date.replace(day=1) + pd.DateOffset(months=1)
                current_date = next_month.to_pydatetime()

        return capex_monthly

    # ----------------------------------------------------
    # Método principal para el flujo completo de CAPEX
    # ----------------------------------------------------
    def fetch_and_distribute_capex(self, year=None):
        """
        Método "todo en uno":
        1) Obtiene datos CAPEX filtrados desde CDF.
        2) Procesa el DataFrame (Start, End, buffer).
        3) Distribuye días por mes.
        4) Retorna un dict con {'January': X, 'February': Y, ...}
        """
        # 1) Obtener DataFrame CAPEX filtrado
        df_capex = self.fetch_capex_activities_for_year(year)
        if df_capex.empty:
            return {}

        # 2) Procesar (convertir fechas, calcular End, etc.)
        df_processed = self.process_capex_data(df_capex)
        if df_processed.empty:
            return {}

        # 3) Distribuir días por mes
        capex_monthly = self.distribute_days_by_month(df_processed)
        return capex_monthly

    # ---------------------------------------------------
    # Métodos para procesar archivos Excel
    # ---------------------------------------------------
    def load_activities_template(self, file_path: str) -> pd.DataFrame:
        """
        Carga la plantilla de actividades-servicios-líneas.
        """
        return pd.read_excel(file_path)
    
    
    def load_catalog_data(self, file_path: str, sheet_name: str = "Bits, Drilling Tools") -> pd.DataFrame:
        """
        Carga el catálogo de costos (Servicios, Línea, Costo) desde la hoja especificada.
        Por defecto se usa la hoja "WCD" para Bits.
        """
        return pd.read_excel(file_path, sheet_name=sheet_name)
    
    def load_excel_file(self):
        """Carga el archivo Excel seleccionado"""
        file_path = r"C:\Users\ncachipuendo\OneDrive - SLB\06 Budget Tool\01 RIG\Approved_Study\CDF_sched_1.7.0.xlsx"
        
        if file_path:
            self.controller.load_excel_file(file_path)
            

    def process_excel_file(self, file_path):
        """
        Carga el archivo Excel, extrae las tablas OPEX y CAPEX y las une con datos de Cognite.
        """
        wb = load_workbook(file_path, data_only=True)
        sheet = wb['Economics Summary']
        
        # Extraer tabla OPEX
        opex_df = self.extract_table(sheet, 'OPEX')
        cognite_df = self.load_from_cognite()
        merged_opex_df = pd.merge(opex_df, cognite_df, left_on='Job', right_on='ID', how='left')

        # Extraer tabla CAPEX
        capex_df = self.extract_table(sheet, 'CAPEX')
        merged_capex_df = pd.merge(capex_df, cognite_df, left_on='Job', right_on='ID', how='left')
        merged_capex_df = self.clean_merged_capex_data(merged_capex_df)

        return merged_opex_df, merged_capex_df

    def extract_table(self, sheet, table_name):
        """
        Extrae una tabla del sheet de Excel y la retorna como DataFrame.
        """
        table = None
        for tbl in sheet._tables.values():
            if tbl.name == table_name:
                table = tbl
                break
        if table is None:
            raise ValueError(f"No se encontró la tabla '{table_name}' en el archivo Excel")
        data = sheet[table.ref]
        rows = [[cell.value for cell in row] for row in data]
        df = pd.DataFrame(rows[1:], columns=rows[0])
        df['Job'] = df['Job'].str.strip()
        return df

    def clean_merged_capex_data(self, df):
        """
        Limpia el DataFrame de CAPEX eliminando filas vacías y filtrando por 'Job name' que contienen 'W'.
        """
        df_filtered = df[df['Job name'].str.contains('W', na=False)]
        df_filtered = df_filtered[df_filtered['Start'].notnull()]
        return df_filtered

    def load_table_from_excel(self, file_path, sheet_name, table_name):
        """
        Carga una tabla específica de un Excel y la retorna como DataFrame.
        """
        wb = load_workbook(file_path, data_only=True)
        sheet = wb[sheet_name]
        if table_name not in sheet.tables:
            raise ValueError(f"Table '{table_name}' not found in sheet '{sheet_name}'.")
        table = sheet.tables[table_name]
        table_range = table.ref
        data = sheet[table_range]
        rows = [[cell.value for cell in row] for row in data]
        df = pd.DataFrame(rows)
        df.columns = df.iloc[0]
        df = df[1:]
        return df

    def load_budget_data_from_excel(self, file_path, sheet_name, year=2025):
        """
        Carga y filtra los datos de presupuesto desde un Excel para un año dado.
        """
        try:
            data = pd.read_excel(file_path, sheet_name=sheet_name)
            #data = data[data['YEAR'] == year]
            #print("Budget data:")
            #print(data)
            return data
        except Exception as e:
            raise ValueError(f"Error al procesar el archivo Excel: {e}")
        
    def load_plan_actividades_from_excel(self, plan_path, sheet_name="Plan2025") -> pd.DataFrame:
        """
        Carga y normaliza el plan anual de actividades desde Excel.

        Args:
            plan_path (str): Ruta al archivo Excel.
            sheet_name (str): Nombre de la hoja con el plan.

        Returns:
            pd.DataFrame: DataFrame con columnas ordenadas y nombres de meses normalizados.
        """
        plan_df = pd.read_excel(plan_path, sheet_name=sheet_name)
        plan_df.fillna(0, inplace=True)
        plan_df.columns = [str(c).strip() for c in plan_df.columns]

        columnas_fijas = ['Tipo de Actividad', 'Total', 'No.']
        columnas_a_normalizar = [col for col in plan_df.columns if col not in columnas_fijas]

        columnas_normalizadas = normalize_month_names(pd.Series(columnas_a_normalizar)).tolist()
        all_months = ["January", "February", "March", "April", "May", "June",
                    "July", "August", "September", "October", "November", "December"]

        new_columns = []
        idx = 0
        for col in plan_df.columns:
            if col in columnas_fijas:
                new_columns.append(col)
            else:
                new_columns.append(columnas_normalizadas[idx])
                idx += 1
        plan_df.columns = new_columns

        for mes in all_months:
            if mes not in plan_df.columns:
                plan_df[mes] = 0

        ordered_columns = ['No.', 'Tipo de Actividad'] + all_months + ['Total']
        plan_df = plan_df[[col for col in ordered_columns if col in plan_df.columns]]

        return plan_df

    def get_total_activities_by_month_df_from_plan(self, plan_path, sheet_name):
        """
        Suma las actividades planificadas por mes usando el plan anual cargado desde Excel.
        Retorna un DataFrame con columnas 'MONTH' y 'PLANNED_ACTIVITIES'.
        """
        activities_df = self.load_plan_actividades_from_excel(plan_path, sheet_name)
        if not activities_df.empty:
            meses = [col for col in activities_df.columns if col not in ['No.', 'Tipo de Actividad', 'Total']]
            monthly_totals = {mes: activities_df[mes].sum() for mes in meses}
            forecast_df = pd.DataFrame({
                'MONTH': list(monthly_totals.keys()),
                'PLANNED_ACTIVITIES': list(monthly_totals.values())
            })
        else:
            months = [calendar.month_name[i] for i in range(1, 13)]
            forecast_df = pd.DataFrame({'MONTH': months, 'PLANNED_ACTIVITIES': [0]*12})  
        return forecast_df

    def load_cotizacion_data(self, file_path):
        """
        Carga y procesa datos de cotización desde un archivo Excel.
        """
        sheet_name = 'Compilado'
        warnings.filterwarnings("ignore")
        wb = load_workbook(file_path, data_only=True)
        ws = wb[sheet_name]
        table = ws.tables['Table2']
        data = ws[table.ref]
        rows = [[cell.value for cell in row] for row in data]
        df = pd.DataFrame(rows[1:], columns=rows[0])
        filtered_df = df[df['OPEX / CAPEX / CPI'] == 'OPEX']
        relevant_columns = ['No', 'POZO', 'MES', 'AÑO', '$ de B&H 2020, 2021']
        filtered_relevant_df = filtered_df[relevant_columns]
        columns_al_to_av = df.loc[:, 'Bomba als (Sarta + misceláneos':'instalacion $']
        columns_al_to_av = columns_al_to_av.apply(pd.to_numeric, errors='coerce')
        columns_al_to_av['Total'] = columns_al_to_av.sum(axis=1)
        final_df = pd.concat([filtered_relevant_df, columns_al_to_av], axis=1)
        print("Cotización data:")
        print(final_df.head())
        return final_df

    # ---------------------------------------------------
    # Métodos para Actividades y Jobs
    # ---------------------------------------------------
    def load_executed_activities(self, year=2024) -> pd.DataFrame:
        """
        Carga las actividades ejecutadas desde el presupuesto anual (filtradas por WO OPEX)
        y las agrupa por mes con nombre estandarizado.

        Returns:
            pd.DataFrame: DataFrame con columnas Month y Executed_Activities.
        """
        df = self.load_budget_data_per_year(year=year)
        if df.empty:
            return pd.DataFrame(columns=["Month", "Executed_Activities"])

        # Agrupar por MONTH (nombre) y contar cantidad de actividades
        actual_activities = df.groupby("MONTH").size().reset_index(name="Executed_Activities")

        # Ordenar por mes calendario
        month_order = get_all_months()
        actual_activities["Month"] = pd.Categorical(actual_activities["MONTH"], categories=month_order, ordered=True)

        # Devolver ordenado
        return actual_activities.sort_values("Month")[["Month", "Executed_Activities"]].reset_index(drop=True)


