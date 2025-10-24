import pandas as pd
from openpyxl import load_workbook

class OpexDataManager:
    """
    Clase encargada de gestionar los datos presupuestarios OPEX (Operational Expenditure).
    Permite cargar, actualizar y guardar los datos presupuestarios asociados a cada línea operativa.

    Atributos:
        data_loader (DataLoader): Componente responsable de acceder a archivos Excel y otras fuentes.
        file_path (str): Ruta del archivo de Excel que contiene la hoja 'OPEX Budget'.
        opex_data (pd.DataFrame): Contenedor interno del DataFrame con los datos OPEX.
    """
    def __init__(self, data_loader, file_path):
        self.data_loader = data_loader
        self.file_path = file_path
        self.opex_data = None

    def load_opex_data(self):
        """
        Carga los datos desde la hoja 'OPEX Budget' utilizando el DataLoader si no se ha cargado previamente.

        Returns:
            pd.DataFrame: DataFrame con columnas 'LINE' y 'OPEX_BUDGET', filtrado por líneas tipo 1.xx.
        """
        if self.opex_data is not None:
            return self.opex_data

        try:
            print(f"📄 Cargando OPEX desde: {self.file_path}")
            df = self.data_loader.load_budget_data_from_excel(self.file_path, sheet_name="OPEX Budget")

            if 'LINE' not in df.columns or 'OPEX_BUDGET' not in df.columns:
                print("❌ Formato inesperado en la hoja 'OPEX Budget'")
                self.opex_data = pd.DataFrame()
                return self.opex_data

            df['OPEX_BUDGET'] = pd.to_numeric(df['OPEX_BUDGET'], errors='coerce').fillna(0.0)
            df = df[df['LINE'].str.contains('1\.')]

            self.opex_data = df.reset_index(drop=True)
            print(f"✅ OPEX cargado correctamente: {len(self.opex_data)} líneas encontradas")
            return self.opex_data

        except Exception as e:
            print(f"❌ Error cargando datos de OPEX: {e}")
            return pd.DataFrame()

    def get_opex_for_line(self, line_name):
        """
        Devuelve el valor presupuestado de OPEX correspondiente a una línea específica.

        Args:
            line_name (str): Nombre de la línea (ej. "1.06 Wireline").

        Returns:
            float: Valor de presupuesto OPEX. Devuelve 0.0 si no se encuentra coincidencia.
        """
        df = self.load_opex_data()

        match = df[df['LINE'].str.strip().str.casefold() == line_name.strip().casefold()]
        if not match.empty:
            return float(match['OPEX_BUDGET'].values[0])
        else:
            print(f"⚠️ No se encontró coincidencia para: {line_name}")
            return 0.0

    def update_opex_for_line(self, line_name, new_value):
        """
        Actualiza el valor presupuestario OPEX para una línea específica.

        Args:
            line_name (str): Nombre de la línea.
            new_value (float): Nuevo valor a asignar.
        """
        df = self.load_opex_data()
        idx = df[df['LINE'].str.contains(line_name.strip(), case=False, na=False)].index
        if not idx.empty:
            self.opex_data.at[idx[0], 'OPEX_BUDGET'] = float(new_value)

    def set_opex_data(self, new_df):
        """
        Reemplaza el DataFrame actual de OPEX con uno nuevo.

        Args:
            new_df (pd.DataFrame): Nuevo DataFrame con columnas 'LINE' y 'OPEX_BUDGET'.
        """
        self.opex_data = new_df.copy()

    def save_opex_to_excel(self):
        """
        Guarda los cambios realizados en los datos OPEX dentro del archivo Excel original.

        La hoja anterior "OPEX Budget" será reemplazada por una nueva hoja con los valores actualizados.
        """
        try:
            wb = load_workbook(self.file_path)

            if 'OPEX Budget' in wb.sheetnames:
                std = wb['OPEX Budget']
                wb.remove(std)

            ws = wb.create_sheet('OPEX Budget')

            headers = ['LINE', 'OPEX_BUDGET']
            for col_idx, header in enumerate(headers, start=1):
                ws.cell(row=1, column=col_idx, value=header)

            for row_idx, row in self.opex_data.iterrows():
                ws.cell(row=row_idx + 2, column=1, value=row['LINE'])
                ws.cell(row=row_idx + 2, column=2, value=row['OPEX_BUDGET'])

            wb.save(self.file_path)
            print(f"✅ Cambios guardados en '{self.file_path}' hoja 'OPEX Budget'")

        except Exception as e:
            print(f"❌ Error al guardar OPEX en Excel: {e}")
